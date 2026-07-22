"""Phase 2 backend: chart detection, anomaly flags, export, derived datasets, dictionary."""
import io
import json

import pytest
from sqlalchemy.orm import Session

from db.models import DatasetRow, RunRow
from graph.anomalies import build_flags
from graph.charts import build_chart_spec


# ---------------------------------------------------------------- charts

def test_month_series_becomes_sorted_line():
    spec = build_chart_spec(
        ["month", "firs"],
        [["2025-03", 30], ["2025-01", 10], ["2025-02", 20]],
    )
    assert spec["type"] == "line"
    assert [p["x"] for p in spec["points"]] == ["2025-01", "2025-02", "2025-03"]
    assert [p["y"] for p in spec["points"]] == [10, 20, 30]


def test_categories_become_bar():
    spec = build_chart_spec(["district", "n"], [["Lucknow", 276], ["Agra", 120], ["Meerut", 90]])
    assert spec["type"] == "bar"
    assert spec["x"] == "district" and spec["y"] == "n"
    assert len(spec["points"]) == 3


def test_single_number_and_wide_results_get_no_chart():
    assert build_chart_spec(["total"], [[3200]]) is None
    assert build_chart_spec(["a", "b", "c"], [[1, 2, 3], [4, 5, 6]]) is None
    assert build_chart_spec(["d", "n"], [["x", 1]]) is None                 # 1 row
    assert build_chart_spec(["d", "n"], [["x", "not-a-number"], ["y", "no"]]) is None


def test_year_axis_counts_as_line():
    spec = build_chart_spec(["year", "n"], [["2024", 1700], ["2025", 1500]])
    assert spec["type"] == "line"


# ---------------------------------------------------------------- flags

def test_month_gap_flagged():
    result = {"rows": [["2025-01", 5], ["2025-02", 7], ["2025-04", 9], ["2025-06", 2]]}
    flags = build_flags(result, [], [], "SELECT ...")
    assert any(f["kind"] == "coverage-gap" for f in flags)
    gap = next(f for f in flags if f["kind"] == "coverage-gap")
    assert "2025-03" in gap["message"] and "2025-05" in gap["message"]


def test_clean_series_no_flags():
    result = {"rows": [["2025-01", 5], ["2025-02", 7], ["2025-03", 9]]}
    assert build_flags(result, [], [], "SELECT ...") == []


def test_high_null_column_flagged_only_when_used():
    datasets = [{
        "id": "d1", "name": "firs", "row_count": 100,
        "columns": [{"name": "crime_head", "null_count": 60}, {"name": "district", "null_count": 0}],
    }]
    used = build_flags({"rows": []}, datasets, ["d1"], "SELECT crime_head, COUNT(*) FROM t GROUP BY crime_head")
    assert any(f["kind"] == "high-nulls" and "crime_head" in f["message"] for f in used)
    unused = build_flags({"rows": []}, datasets, ["d1"], "SELECT district, COUNT(*) FROM t GROUP BY district")
    assert not any(f["kind"] == "high-nulls" for f in unused)


# ---------------------------------------------------------------- export

@pytest.fixture
def completed_run(api_client, _isolated_db):
    from tests_helpers import read_sample
    r = api_client.post("/datasets", files=[("files", ("fir_records.csv", read_sample("fir_records.csv"), "text/csv"))])
    table = r.json()["data"][0]["table_name"]
    with Session(_isolated_db) as s:
        run = RunRow(
            status="completed",
            input_text="All Lucknow FIRs",
            sql_text=f"SELECT fir_no, district, fir_date FROM \"{table}\" WHERE district='Lucknow'",
        )
        s.add(run)
        s.commit()
        return run.id


def test_export_csv_returns_full_rows_beyond_preview_cap(api_client, completed_run, expected):
    r = api_client.get(f"/runs/{completed_run}/export?format=csv")
    assert r.status_code == 200
    assert "attachment" in r.headers["content-disposition"]
    lines = [ln for ln in r.text.splitlines() if ln.strip()]
    lucknow_total = sum(1 for _ in lines) - 1  # minus header
    assert lucknow_total > 200, "export must exceed the 200-row preview cap (full data)"
    assert lucknow_total >= expected["fir_2025_by_district"]["Lucknow"]


def test_export_xlsx_has_data_and_about_sheets(api_client, completed_run):
    r = api_client.get(f"/runs/{completed_run}/export?format=xlsx")
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    assert set(wb.sheetnames) == {"Data", "About"}


def test_export_after_dataset_deleted_409(api_client, completed_run):
    ds = api_client.get("/datasets").json()["data"][0]
    api_client.delete(f"/datasets/{ds['id']}")
    r = api_client.get(f"/runs/{completed_run}/export?format=csv")
    assert r.status_code == 409
    assert "deleted" in r.json()["detail"]["message"]


def test_export_bad_format_and_unknown_run(api_client, completed_run):
    assert api_client.get(f"/runs/{completed_run}/export?format=pdf").status_code == 400
    assert api_client.get("/runs/nope/export").status_code == 404


# ---------------------------------------------------------------- derived datasets

def test_derived_dataset_saved_and_queryable(api_client, completed_run):
    r = api_client.post("/datasets/derived", json={"run_id": completed_run, "name": "Lucknow FIRs"})
    assert r.status_code == 200, r.text
    ds = r.json()["data"]
    assert ds["source"] == "derived"
    assert ds["row_count"] > 200
    from ingest import store
    check = store.run_select(f'SELECT COUNT(*) FROM "{ds["table_name"]}"')
    assert check["rows"][0][0] == ds["row_count"]
    prof = api_client.get("/datasets").json()["data"]
    me = next(d for d in prof if d["id"] == ds["id"])
    assert me["profile"]["provenance"]["run_id"] == completed_run


def test_derived_fails_honestly_when_source_deleted(api_client, completed_run):
    src = api_client.get("/datasets").json()["data"][0]
    api_client.delete(f"/datasets/{src['id']}")
    r = api_client.post("/datasets/derived", json={"run_id": completed_run, "name": "x"})
    assert r.status_code == 409


# ---------------------------------------------------------------- dictionary

def test_dictionary_edit_persists_and_reaches_prompt_catalog(api_client, _isolated_db):
    r = api_client.post("/datasets", files=[("files", ("p.csv", b"district,crime_head\nLucknow,Theft\nAgra,Dowry Harassment\n", "text/csv"))])
    ds = r.json()["data"][0]
    patch = api_client.patch(
        f"/datasets/{ds['id']}/columns/crime_head",
        json={"description": "CCTNS head codes; POCSO grouped under Kidnapping"},
    )
    assert patch.status_code == 200
    cols = {c["name"]: c for c in patch.json()["data"]["columns"]}
    assert "CCTNS" in cols["crime_head"]["description"]

    # survives a fresh read AND appears in the catalog text the agent prompts with
    listed = api_client.get("/datasets").json()["data"][0]
    assert "CCTNS" in {c["name"]: c for c in listed["columns"]}["crime_head"]["description"]

    from graph.nodes import _dataset_catalog
    catalog = _dataset_catalog(
        [{"id": ds["id"], "table_name": ds["table_name"], "name": ds["name"],
          "row_count": 2, "columns": listed["columns"]}],
        full=True,
    )
    assert "CCTNS" in catalog

    assert api_client.patch(f"/datasets/{ds['id']}/columns/nope", json={"description": "x"}).status_code == 404
    assert api_client.patch("/datasets/nope/columns/crime_head", json={"description": "x"}).status_code == 404
